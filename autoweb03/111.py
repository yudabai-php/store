from openpyxl import load_workbook


write = load_workbook(filename='./HKR.xlsx')
print(write.sheetnames)

sheet = write.active

call = sheet['H3']
call.value = '白告子'

write.save(filename='./HKR.xlsx')